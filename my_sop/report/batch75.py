# -*- coding: utf-8 -*-
import numpy as np
import sys
from os.path import abspath, join, dirname
from datetime import datetime

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
sys.path.insert(0, join(abspath(dirname(__file__)), '../'))

from sop.connect_clickhouse import connect


def get_shopmap(filename):
    df = pd.read_excel(filename,header=0,usecols=['Store','Sub-Channel','Store (Grouped)'],sheet_name='storename_map')[['Store','Sub-Channel','Store (Grouped)']]
    data = np.array(df)
    map = {}
    for d in data:
        m = {}
        m['channels']=d[1]
        m['storegroups']=d[2]
        map[d[0]]=m
    return map

def get_sql_info(transform_channels,transform_storegroups,where):
    sql = """
            SELECT YEAR(pkey) "year"
            ,IF(spmodality = 'Others','CCL',spmodality) "modality" 
            ,`sp散光老花`
            ,if(item_id = '775428806073','透明片',spPlatform) category
            ,transform(alias_all_bid,[6071652,52390,243393,311028,201237,176723,130598,109463,7687673,6153805,4846759,243071,7684675,2879494,6343661,6474588],['J&J','J&J','B+L','COOPER VISION','ALCON','HORIEN','HYDRON','WEICON','YOUTO','YOUTO','T-Garden','KILALA','KILALA','MOODY','LAPECHE','LAPECHE'],'Other-other') AS "Manufacturer China"
            ,{transform_channels}
            ,{transform_storegroups}
            ,SUM(sales)/100 "销售额"
            FROM sop_e.entity_prod_90613_E e
            {where}
            GROUP BY "year","modality",`sp散光老花`,category,"Manufacturer China",channels,storegroups
        """.format(transform_channels = transform_channels,transform_storegroups=transform_storegroups,where=where)
    return sql

def get_data(map,where):
    shop_name_list = list(map.keys())
    channels_list = [map[s]['channels'] for s in shop_name_list]
    storegroups_list = [map[s]['storegroups'] for s in shop_name_list]
    transform_channels = """ transform(dictGet('all_shop','nick',tuple(toUInt8(`source`),toUInt32(e.sid))),{shop_name_list},{channels_list},'') AS channels """.format(shop_name_list=shop_name_list,channels_list=channels_list)
    transform_storegroups = """ transform(dictGet('all_shop','nick',tuple(toUInt8(`source`),toUInt32(e.sid))),{shop_name_list},{storegroups_list},'') AS storegroups """.format(shop_name_list=shop_name_list, storegroups_list=storegroups_list)
    sql = get_sql_info(transform_channels=transform_channels,transform_storegroups=transform_storegroups,where=where)
    df = connect(0,sql)
    return df

def write_xlsx(data,row_s,col_s,work_book,sheet_name):
    sheet = work_book[sheet_name]
    for r in range(len(data)):
        for c in range(len(data[0])):
            sheet.cell(row_s+r,col_s+c).value = data[r][c]
    return work_book

def run1(work_book):
    category_map = {'Clear':'透明片','Beauty':'彩片','CS':'透明片'}
    rows = ['Clear','CS DD','CS RD','CS CCL','toric','multifocal','Beauty','Beauty DD','Beauty RD','Beauty CCL','Category']
    cols = ['Total Category Sales Value(RMB)','J&J Sales Value(RMB)']
    year = [2023,2024]
    map = get_shopmap(map_filename)
    res = get_data(map, where)
    table1 = []
    table2 = []
    for r in rows:
        d1,d2 = [],[]
        for c in cols:
            # 区分是否强生
            if c == 'J&J Sales Value(RMB)':
                res_tem = res.loc[res["Manufacturer China"]=='J&J']
            else:
                res_tem = res.copy()
            # 以rows来限制不同的取数条件
            if r in ['Clear','Beauty']:
                data = res_tem.loc[res_tem['category']==category_map[r]]
            if r in ['Category']:
                data = res_tem.copy()
            if r in ['toric','multifocal']:
                data = res_tem.loc[(res_tem['sp散光老花']==r) & (res_tem['category']=='透明片')]
            if r in ['CS DD','CS RD','CS CCL','Beauty DD','Beauty RD','Beauty CCL']:
                rule = r.split(' ')
                if len(rule)==2:
                    data = res_tem.loc[(res_tem['category']==category_map[rule[0]]) & (res_tem['modality']==rule[1])]
            if c == 'J&J Sales Value(RMB)' and r in ['CS CCL','Beauty CCL','multifocal']:
                d1 += ['','','']
                d2 += ['','','']
            else:
                d1 += [data.loc[data['year'] == year[0]]["销售额"].sum(), data.loc[data['year'] == year[1]]["销售额"].sum(),data.loc[data['year'] == year[1]]["销售额"].sum() / data.loc[data['year'] == year[0]]["销售额"].sum() - 1]
                d2 += [data.loc[data['year'] == year[0]]["销售额"].sum() / res_tem.loc[res_tem['year'] == year[0]]["销售额"].sum(),data.loc[data['year'] == year[1]]["销售额"].sum() / res_tem.loc[res_tem['year'] == year[0]]["销售额"].sum(), (data.loc[data['year'] == year[1]]["销售额"].sum() -data.loc[data['year'] == year[0]]["销售额"].sum()) /res_tem.loc[res_tem['year'] == year[0]]["销售额"].sum()]
        table1.append(d1)
        table2.append(d2)
    work_book = write_xlsx(table1, 5, 2, work_book, 'Seg-modality')
    work_book = write_xlsx(table2, 19, 2, work_book, 'Seg-modality')
    return work_book

def run2(work_book):
    rows = ['Total','Beauty','Beauty DD','Beauty RD', 'Clear', 'CS DD','CS RD']
    manufacturer = ['J&J', 'B+L', 'COOPER VISION', 'ALCON', 'HYDRON', 'HORIEN', 'WEICON', 'OTHER MANUFACTURERS','YOUTO', 'T-Garden', 'KILALA', 'MOODY', 'LAPECHE', 'Other-other']
    map = get_shopmap(map_filename)
    res = get_data(map, where)
    table1 = []
    for r in rows:
        res_tem = res.copy()
        if r in ['Clear', 'Beauty']:
            res_tem = res_tem.loc[res_tem['category'] == category_map[r]]
        if r in ['Total']:
            res_tem = res_tem.copy()
        if r in ['CS DD', 'CS RD', 'Beauty DD', 'Beauty RD']:
            rule = r.split(' ')
            if len(rule) == 2:
                res_tem = res_tem.loc[(res_tem['category'] == category_map[rule[0]]) & (res_tem['modality'] == rule[1])]

        ttl_sales1 = res_tem.loc[res_tem['year'] == year[0]]["销售额"].sum()
        ttl_sales2 = res_tem.loc[res_tem['year'] == year[1]]["销售额"].sum()

        for m in manufacturer:
            if m == 'OTHER MANUFACTURERS':
                data = res_tem.loc[res_tem["Manufacturer China"].isin(['YOUTO', 'T-Garden', 'KILALA', 'MOODY', 'LAPECHE', 'Other-other'])]
            else:
                data = res_tem.loc[res_tem["Manufacturer China"]==m]
            if data.loc[data['year'] == year[0]]["销售额"].sum() != 0:
                d1 = [m,data.loc[data['year'] == year[0]]["销售额"].sum(), data.loc[data['year'] == year[1]]["销售额"].sum(),data.loc[data['year'] == year[1]]["销售额"].sum() / data.loc[data['year'] == year[0]]["销售额"].sum() - 1]
            else:
                d1 = [m,data.loc[data['year'] == year[0]]["销售额"].sum(), data.loc[data['year'] == year[1]]["销售额"].sum(),0]
            d2 = [data.loc[data['year'] == year[0]]["销售额"].sum()/ttl_sales1, data.loc[data['year'] == year[1]]["销售额"].sum()/ttl_sales2,data.loc[data['year'] == year[1]]["销售额"].sum()/ttl_sales1-data.loc[data['year'] == year[0]]["销售额"].sum()/ttl_sales2]
            d = d1+d2
            table1.append(d)
        table1.append(['','','','','','',''])
    work_book = write_xlsx(table1, 34, 2, work_book, 'Seg-modality')
    return work_book

def run3(work_book):
    rows = ['DTC','T6 Flagship Stores','Non-T6 Flagship Stores','EKA','Core Stores','Exclusive Store','Platform Store','Ali Health','TDI','Tmall Supermarket','E-DISTRIBUTION','Domestic Trade','Overseas Trade','TOTAL CATEGORY']
    cols = ['Total Category Sales Value(RMB)','J&J Sales Value(RMB)']
    year = [2023,2024]
    map = get_shopmap(map_filename)
    res = get_data(map, where)
    table1 = []
    for r in rows:
        d1,d2 = [],[]
        for c in cols:
            # 区分是否强生
            if c == 'J&J Sales Value(RMB)':
                res_tem = res.loc[res["Manufacturer China"]=='J&J']
            else:
                res_tem = res.copy()
            # 以rows来限制不同的取数条件
            if r in ['DTC','EKA','Platform Store','E-DISTRIBUTION']:
                data = res_tem.loc[res_tem['channels']==r]
            if r in ['T6 Flagship stores','Non-T6 Flagship Stores','Core Stores','Exclusive Store','Ali Health','TDI','Tmall supermarket','Domestic Trade','Overseas Trade']:
                data = res_tem.loc[res_tem['storegroups'] == r]
            if r in ['TOTAL CATEGORY']:
                data = res_tem.copy()

            d1 += [data.loc[data['year'] == year[0]]["销售额"].sum(), data.loc[data['year'] == year[1]]["销售额"].sum(),data.loc[data['year'] == year[1]]["销售额"].sum() / data.loc[data['year'] == year[0]]["销售额"].sum() - 1]
        table1.append(d1)
    work_book = write_xlsx(table1, 8, 2, work_book, 'Seg-channel')
    return work_book

def run4(work_book):
    rows = ['DTC','EKA','Platform Store','DTC+EKA', 'Edistri', 'T6 Flagship Stores','Ali health/Tmall supermarket']
    year = [2023, 2024]
    manufacturer = ['J&J', 'B+L', 'COOPER VISION', 'ALCON', 'HYDRON', 'HORIEN', 'WEICON', 'OTHER MANUFACTURERS','YOUTO', 'T-Garden', 'KILALA', 'MOODY', 'LAPECHE', 'Other-other']
    map = get_shopmap(map_filename)
    res = get_data(map, where)
    table1 = []
    for r in rows:
        res_tem = res.copy()
        if r in ['DTC','EKA','Platform Store', 'Edistri']:
            res_tem = res_tem.loc[res_tem['channels'] == r]
        if r in ['T6 Flagship Stores']:
            res_tem = res_tem.loc[res_tem['storegroups'] == r]
        if r in ['DTC+EKA']:
            res_tem = res_tem.loc[res_tem['channels'].isin(r.split('+'))]
        if r in ['Ali health/Tmall supermarket']:
            res_tem = res_tem.loc[res_tem['storegroups'].isin(r.split('/'))]
        ttl_sales1 = res_tem.loc[res_tem['year'] == year[0]]["销售额"].sum()
        ttl_sales2 = res_tem.loc[res_tem['year'] == year[1]]["销售额"].sum()

        for m in manufacturer:
            if m == 'OTHER MANUFACTURERS':
                data = res_tem.loc[res_tem["Manufacturer China"].isin(['YOUTO', 'T-Garden', 'KILALA', 'MOODY', 'LAPECHE', 'Other-other'])]
            else:
                data = res_tem.loc[res_tem["Manufacturer China"]==m]
            if data.loc[data['year'] == year[0]]["销售额"].sum() != 0:
                d1 = [m,data.loc[data['year'] == year[0]]["销售额"].sum(), data.loc[data['year'] == year[1]]["销售额"].sum(),data.loc[data['year'] == year[1]]["销售额"].sum() / data.loc[data['year'] == year[0]]["销售额"].sum() - 1]
            else:
                d1 = [m,data.loc[data['year'] == year[0]]["销售额"].sum(), data.loc[data['year'] == year[1]]["销售额"].sum(),0]
            d2 = [data.loc[data['year'] == year[0]]["销售额"].sum()/ttl_sales1, data.loc[data['year'] == year[1]]["销售额"].sum()/ttl_sales2,data.loc[data['year'] == year[1]]["销售额"].sum()/ttl_sales1-data.loc[data['year'] == year[0]]["销售额"].sum()/ttl_sales2]
            d = d1+d2
            table1.append(d)
        table1.append(['','','','','','',''])
    work_book = write_xlsx(table1, 28, 3, work_book, 'Seg-channel')
    return work_book

def run(work_book):
    work_book = run1(work_book)
    work_book = run2(work_book)
    work_book = run3(work_book)
    work_book = run4(work_book)
    return work_book

if __name__ == "__main__":
    global work_book,year,category_map
    project_path = '../media/batch75/'
    template = 'Tmall store mapping_whole version_0701.xlsx'
    table_template = '2024.6.18 topline structure for J&J_240607.xlsx'
    output_name = '2024.6.18 topline structure for J&J_{}.xlsx'.format(datetime.now().strftime('%y%m%d'))
    year = [2023, 2024]
    category_map = {'Clear': '透明片', 'Beauty': '彩片', 'CS': '透明片'}
    where = """ 
            WHERE `source` = 1
            AND spPlatform in ('透明片','彩片')
            AND ((`date` >= '2024-05-20'AND `date` < '2024-06-21') or (`date` >= '2023-05-20'AND `date` < '2023-06-21'))
            """
    map_filename = project_path + template
    table_filename = project_path + table_template
    work_book = load_workbook(project_path + table_template)
    output = run(work_book)
    output.save(project_path + output_name)



