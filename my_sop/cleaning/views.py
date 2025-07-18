import datetime
import time
import ast
from django.shortcuts import render,get_object_or_404
from openpyxl import load_workbook
from io import BytesIO
from .models import CleanBatchLog,CleanBatch,CleanCron
from sop import connect_clickhouse
from django.http import JsonResponse,FileResponse
# from django.db.models import Max
from django.views.decorators.csrf import csrf_exempt
# from django.conf import settings
import os
import redis
# from django.core.paginator import Paginator
import mimetypes
from urllib.parse import quote
import pandas as pd
import json
import importlib
import sys
import django_rq
from django.db import connection
from django.db.models import Q
from os.path import abspath, join, dirname
sys.path.insert(0, join(abspath(dirname(__file__)), ''))
# import application as app
# import models.entity_manager as entity_manager

def dynamic_import(batch_id):
    try:
        module = importlib.import_module("cleaning.model.plugins.batch{}.run".format(batch_id), package=__name__)
        return module
    except ImportError as e:
        # 处理导入错误
        print(f"Error importing module: {e}")
        return None

def cleaner(request):
    if request.method == 'GET':
        batchId = request.GET.get('id')
        batchName = CleanBatch.objects.get(batch_id=batchId).name
        eid = CleanBatch.objects.get(batch_id=batchId).eid
        print(batchId,batchName,eid)
        return render(request, 'cleaning/cleaning-flow.html', {"batchId":batchId,"batchName":batchName,"eid":eid})

    return render(request, 'sop/404.html', locals())

def share_rules(request,id):
    file_path = r'./cleaning/model/plugins/batch{}/rules/rules.xlsx'.format(id)
    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'))  # 打开文件进行读取
        mime_type, _ = mimetypes.guess_type(file_path)
        response['Content-Type'] = mime_type or 'application/octet-stream'
        filename = os.path.basename(file_path)
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(quote(filename))  # 设置为附件形式并指定默认文件名
        return response
    else:
        # 如果文件不存在，返回404
        return JsonResponse({'code':404,'errmsg':'rules配置文件未找到。'})

def clean_rules(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        batchId = request.POST.get('batchId')
        FileName = request.POST.get('FileName')
        if file:
            # 读取 Excel 文件
            file_data = file.read()
            workbook = load_workbook(filename=BytesIO(file_data))
            workbook.save(r'./cleaning/model/plugins/batch{}/rules/rules.xlsx'.format(batchId, batchId))
            queue = django_rq.get_queue('default')
            job = queue.enqueue(rules_save, batchId)
            print(job.get_status())
            # result = rules_save.apply_async((batchId,), queue='default_queue')
            # result = AsyncResult(result.task_id)
            # if result.ready():
            #     return JsonResponse({'status': 'completed', 'result': result.result})
            # else:
            #     return JsonResponse({'status': 'pending'})
            js = {'code': 200,
                  'data': '文件上传成功\n共包含【{}】等[{}]个sheet\ntask_id:{}'.format(workbook.sheetnames[0],len(workbook.sheetnames),job.id)}
            if int(batchId) == 362:
                module = dynamic_import(batchId)
                if module:
                    print("{}开始convert_brand".format(batchId))
                    ret = module.convert_brand(job)
                    js['data'] += 'convert_brand更新状态：{}'.format(ret)
            return JsonResponse(js)
        else:
            return JsonResponse({'code': -1, 'errMsg': '文件上传失败'})
    BatchID = request.GET.get('batchId')
    convert_brand = 'None'
    if int(BatchID) == 362:
        convert_brand = 'black'
    return render(request, 'cleaning/clean_rules.html', {"BatchID":BatchID,"display":convert_brand})

def download_rules(request):
    batchId = request.GET.get('batchId')
    file = request.GET.get('file','rules')
    print(file)
    file_path = os.path.join(f'./cleaning/model/plugins/batch{batchId}/rules/{file}.xlsx')  # 构建文件的绝对路径
    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'))  # 打开文件进行读取
        mime_type, _ = mimetypes.guess_type(file_path)
        response['Content-Type'] = mime_type or 'application/octet-stream'
        filename = os.path.basename(file_path)
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(quote(filename))# 设置为附件形式并指定默认文件名
        return response
    else:
        # 如果文件不存在，返回404
        return render(request, 'sop/404.html',locals())

def clean_process(request):
    if request.method == 'POST':
        batchId = request.POST.get('batchId')
        uuid2 = request.POST.get('uuid')
        define = request.POST.get('define')
        if define:
            define_json = json.loads(define)
        else:
            define_json = {}
        print(define_json)
        module = dynamic_import(batchId)
        if module:
            print("模型导入")
            process_log = module.process_log(uuid2,define_json)
            print({'code': 200, "data": process_log})
            return JsonResponse({'code': 200, "data": process_log})
        print("batch{}查询uuid2{}".format(batchId,uuid2))
        return JsonResponse({'code': -1, 'errmsg': '清洗配置未找到'})
    return render(request, 'cleaning/clean_process.html', locals())

def clean_statu(request):
    items = CleanBatchLog.objects.all()
    operation_dict = {
        "导清洗表": [{"id":"syncAtbl","class":"btn btn-info","name":"导清洗表"}],
        "clean": [{"id":"importClean","class":"btn btn-info","name":"添加机洗"}],
        "出题": [{"id":"brush","class":"fold","name":""}],
        "刷答题默认值": [{"id":"updateBrushDefault","class":"btn btn-info","name":"刷答题默认值"}],
        "刷sku默认值": [{"id":"updateSkuDefault","class":"btn btn-info","name":"刷sku默认值"}],
        "更新正确率": [{"id":"updateCorrect","class":"btn btn-info","name":"更新正确率"}],
        "出数到E表": [{"id":"importMarket","class":"btn btn-info","name":"出数"},{"id":"version","class":"btn btn-info","name":"版本记录"}]
    }
    res = []

    for operation in operation_dict.keys():
        found = False
        for item in items:
            if item.type == operation:
                res.append({
                    "operations": operation_dict[operation],
                    "T": item.type,
                    "tips": item.params,
                    "status": item.status,
                    "msg": item.msg,
                    "warn": item.warn,
                    "process": item.process,
                    "create_time": item.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                    "update_time": item.update_time.strftime('%Y-%m-%d %H:%M:%S'),
                })
                found = True
                break

        if not found:
            res.append({
                "operations": operation_dict[operation],
                "T": None,
                "tips": None,
                "status": None,
                "msg": None,
                "warn": None,
                "process": None,
                "create_time": None,
                "update_time": None,
            })

    return JsonResponse({"items": res})

def clean_task(request,id=0):
    def format_time(t):
        if not t:
            return ''

        h = int(t / 3600)
        t = t % 3600
        m = int(t / 60)
        s = int(t % 60)
        return '{}{}{}{}{}s'.format(h, 'h' if h > 0 else '', m, 'm' if m > 0 else '', s)
    if request.method == 'POST':

        try:
            batches = CleanBatch.objects.values('batch_id', 'name')
            bna = {batch['batch_id']: batch['name'] for batch in batches}

            sql = '''
                    SELECT count(DISTINCT server_ip), sum(unix_timestamp(completedTime)-unix_timestamp(beginTime))/count(*)
                    FROM clean_cron WHERE status = 'completed' AND completedTime > date_sub(now(),interval 1 day)
                '''
            with connection.cursor() as cursor:
                cursor.execute(sql)
                ret = cursor.fetchall()
            # 最近1天内子任务平均清洗时间
            server_count, hash_rate = ret[0][0] or 1, ret[0][1] or 0

            # 最近1天内每个batch的子任务平均清洗时间
            sql = '''
                    SELECT batch_id, sum(unix_timestamp(completedTime)-unix_timestamp(beginTime))/count(*)
                    FROM clean_cron WHERE status = 'completed' AND completedTime > date_sub(now(),interval 1 day)
                    GROUP BY batch_id
                '''
            with connection.cursor() as cursor:
                cursor.execute(sql)
                ret = cursor.fetchall()
            his = {bid: avgtime for bid, avgtime, in ret}

            sql = """
                    SELECT * FROM (
                        SELECT
                            batch_id, task_id, cln_tbl,
                            max(emergency) as killed,
                            max(priority) as pri,
                            min(IF(beginTime!='1970-01-01 00:00:00',0,UNIX_TIMESTAMP(planTime))) as plan_time,
                            min(IF(beginTime!='1970-01-01 00:00:00', beginTime, '~')) as begin_time,
                            max(IF(completedTime!='1970-01-01 00:00:00', beginTime, '~')) as comp_time,
                            count(*) as task_count,
                            sum(IF(status='completed',1,0)) as comp_count,
                            sum(IF(status='error',1,0)) as err_count,
                            sum(IF(status='process',1,0)) as proc_count,
                            sum(IF(status='completed', unix_timestamp(completedTime)-unix_timestamp(beginTime),0)) as time_per_task,
                            sum(IF(status NOT IN ('error', 'completed'),`count`,0)) as count_less_task,
                            unix_timestamp(max(beginTime)) as unix_begin_time,
                            min(params) as min_params, max(params) as max_params,
                            min(createTime) as createTime
                        FROM clean_cron
                        WHERE task_id IN (
                            SELECT max(task_id) FROM clean_cron WHERE createTime > date_add(now(), INTERVAL -1 MONTH)
                            GROUP BY eid, cln_tbl, planTime
                        )
                        GROUP BY batch_id,task_id,cln_tbl
                    ) a ORDER BY (task_count>comp_count) DESC, killed, err_count=0 DESC, begin_time DESC
                """

            # 使用 Django 的 raw() 方法执行原生 SQL 查询
            with connection.cursor() as cursor:
                cursor.execute(sql)
                ret = cursor.fetchall()

            ret = list(ret)

            data, tmp_time = [], time.time()
            for i, v in enumerate(ret):
                bid, task_id, cln_tbl, stop, priority, plan, stime, etime, total, comp, fail, pros, t, count, btime, params1, params2, created, = v
                if comp > 0:
                    avgtime = t / comp
                elif bid in his:
                    avgtime = his[bid]
                else:
                    avgtime = hash_rate

                if total != comp + fail:
                    lesstime = avgtime * max(1, (total - comp - fail) / server_count)
                else:
                    lesstime = ''

                if total == comp + fail:
                    # 任务已完成或失败
                    beginTime = ''
                    completedTime = ''
                elif comp > 0 or pros > 0 or fail > 0:
                    # 任务已经开始
                    beginTime = ''
                    completedTime = btime + int(lesstime)
                elif plan > tmp_time:
                    # 设定时间未开始
                    beginTime = plan
                    completedTime = beginTime + int(lesstime)
                    lesstime = ''
                    tmp_time = completedTime
                else:
                    # 即时任务
                    beginTime = tmp_time
                    completedTime = beginTime + int(lesstime)
                    tmp_time = completedTime

                if stop:
                    lesstime = '已终止' if pros == 0 else '终止中'
                    beginTime = '已终止' if pros == 0 else '终止中'
                    completedTime = '已终止' if pros == 0 else '终止中'
                    stime = '已终止' if pros == 0 else '终止中'
                    etime = '已终止' if pros == 0 else '终止中'
                elif total == comp:
                    lesstime = '已完成'
                elif fail > 0:
                    lesstime = '有任务失败了'
                elif beginTime != '':
                    lesstime = '未开始'
                else:
                    lesstime = format_time(lesstime)

                if beginTime and beginTime != '已终止' and beginTime != '终止中':
                    beginTime = datetime.datetime.fromtimestamp(int(beginTime))
                if completedTime and completedTime != '已终止' and completedTime != '终止中':
                    completedTime = datetime.datetime.fromtimestamp(int(completedTime))
                if plan:
                    plan = datetime.datetime.fromtimestamp(int(plan))

                name = bna[bid] if bid in bna else ''
                btn = '<button class="btn btn-info" style="margin-right: 8px;" onclick="window.detail({},{});">查看</button>'.format(
                    bid, task_id)
                if total != comp:
                    btn += '<button class="btn btn-success" style="margin-right: 8px;" onclick="window.modify({},{});">保存</button>'.format(
                        bid, task_id)
                if total != comp + fail and not stop:
                    btn += '<button class="btn btn-danger" style="margin-right: 8px;" onclick="window.stop({},{});">终止</button>'.format(
                        bid, task_id)
                if fail > 0 or stop:
                    btn += '<button class="btn btn-warning" style="margin-right: 8px;" onclick="window.retry({},{});">重试</button>'.format(
                        bid, task_id)

                where = '', '', ''
                if params1 != '':
                    params1 = json.loads(params1)
                    params2 = json.loads(params2)
                    where = params2['w'] if 'w' in params2 and params2['w'] != '' else where

                data.append([
                    bid, name, task_id, cln_tbl, where,
                    int(total) or '', int(comp) or '', int(fail) or '', int(pros) or '',
                    format_time(avgtime), lesstime,
                    str(created),
                    priority, str(plan or ''), str(beginTime or ''), str(completedTime or ''), str(stime or ''),
                    str(etime or ''),
                    btn
                ])

            return JsonResponse({'code': 0, 'data': {'server_count': server_count, 'hash_rate': hash_rate, 'data': data}})

        except Exception as e:
            print(e)
            return JsonResponse({'code': 1, 'data': 'False'})
    else:
        BatchID = request.GET.get('batchId')
        convert_brand = 'None'
        if int(BatchID) in [362]:
            display = 'black'
        else:
            display = 'None'
    return render(request, 'cleaning/clean-task1.html', locals())

def get_connected_clients():
    redis_host = '10.21.90.130'
    redis_port = 6379
    redis_password = 'nint'
    r = redis.Redis(host=redis_host, port=redis_port, password=redis_password, decode_responses=True)
    clients = r.client_list()
    unique_ips = set()
    for client in clients:
        ip = client['addr'].replace('[::1]', '127.0.0.1').split(':')[0]
        unique_ips.add(ip)

    return unique_ips


def view_clean_task(request, id):
    if request.method != 'POST':
        return JsonResponse({'code': 1, 'data': 'Invalid request method'})

    try:
        # 解析和验证输入数据
        data = json.loads(request.POST.get('data', '{}'))
        bid = request.POST.get('id')
        tid = data.get('tid')

        if not bid or not tid:
            return JsonResponse({'code': 1, 'data': 'Missing batch_id or task_id'})

        # 查询数据库并获取指定字段
        ret = CleanCron.objects.filter(batch_id=bid, task_id=tid).order_by('id').values_list(
            'id', 'task_id', 'batch_id', 'eid', 'server_ip', 'process_id', 'priority',
            'retry', 'minCPU', 'minRAM', 'count', 'params', 'status', 'emergency',
            'msg', 'planTime', 'beginTime', 'completedTime', 'createTime'
        )

        # 将查询结果转换为列表
        res = list(ret)

    except json.JSONDecodeError:
        return JsonResponse({'code': 1, 'data': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'code': 1, 'data': str(e)})

    # 打印结果用于调试
    print(res)

    return JsonResponse({'code': 0, 'data': res})

def clean_task_plan(request,id):
    print(request.POST)
    js = {"code": 0, "data": [[]]}
    return JsonResponse(js)


from my_sop.task import rules_save
from celery.result import AsyncResult
def create_task(request):
    # result = my_task.delay('测试任务调度')
    result = my_task.apply_async(('测试任务调度{}'.format(1+1),), queue='default_queue')
    return JsonResponse({'task_id': result.id})


def get_task_result(request, task_id):
    result = AsyncResult(task_id)
    if result.ready():
        return JsonResponse({'status': 'completed', 'result': result.result})
    else:
        return JsonResponse({'status': 'pending'})

def create_clean_task(request,id):
    if request.method == 'POST':
        print(request.POST)
        try:
            data = json.loads(request.POST.get('data'))
        except:
            data = request.POST
        user = data.get('r')
        batch_id = data.get('bid') or data.get('batchId')
        batch_name = data.get('batchName')
        where = data.get('where') or '1'
        priority = data.get('priority',0)
        tbl = data.get('tbl',None)
        s_date = data.get('smonth').replace('/', '-')[0:10]
        e_date = data.get('emonth').replace('/', '-')[0:10]
        is_test = data.get('is_test','')
        test_choose = data.get('test_choose','')
        real_batchName = CleanBatch.objects.get(batch_id=batch_id).name
        eid = CleanBatch.objects.get(batch_id=batch_id).eid
        task_id = int(datetime.datetime.now().timestamp())

        if real_batchName != batch_name:
            js = {
                "code": 300,
                "errmsg": "Batch ID: {batchId} \n正确名称为: {real_batchName} \n输入为: {batchName}\n"
                .format(batchId=batch_id, real_batchName=real_batchName, batchName=batch_name)
            }
            return JsonResponse(js)
        check = len(CleanCron.objects.filter(Q(emergency=0) & Q(batch_id=batch_id) & ~Q(status__in=['completed', 'error']))) \
                + len(CleanCron.objects.filter(Q(emergency=1) & Q(status='process') & Q(batch_id=batch_id) & ~Q(status__in=['completed', 'error'])))
        print("清洗状态检查：",check)
        if check:
            js = {
                "code": 300,
                "errmsg": "batchId{}正在清洗中，请不要重复添加".format(batch_id)
            }
            return JsonResponse(js)

        queue = django_rq.get_queue('default')
        module = dynamic_import(batch_id)
        if module:
            print("batchId{}模型导入".format(batch_id))
            params = json.dumps({
                "w": f"({where}) AND (date >= '{s_date}') AND (date < '{e_date}')",
                "s": s_date,
                "e": e_date
            })
            print("params测试:", params)
            CleanBatchLog.objects.create(batch_id=get_object_or_404(CleanBatch, batch_id=batch_id), eid=eid,
                                         comments=is_test,type='clean', task_id=task_id, status='process', process='清洗中',params=params)
            scripts = module.add_task(batch_id,eid,task_id,priority,test_choose,params)

            job = queue.enqueue(module.cleaning, batch_id,task_id,scripts)
            js = {
                "code": 200,
                "msg": "batchId{}清洗任务添加成功".format(batch_id)
            }
            return JsonResponse(js)
        js = {
            "code": 404,
            "msg": "batchId{}清洗清洗模型异常".format(batch_id)
        }
        return JsonResponse(js)

    else:
        return render(request, 'sop/404.html')

def modify_clean_task(request, id=0):
    print('<modify_clean_task>')
    if request.method == 'POST':
        data = json.loads(request.POST.get('data'))
        print(data)
        tid = data.get('tid', 0)
        pri = int(data.get('pri', 0))
        pt = data.get('pt', '').strip()
        if not pt:
            pt = datetime.datetime(1970, 1, 1, 0, 0)
    print(tid,pri,pt)
    try:
        CleanCron.objects.filter(batch_id=id, task_id=tid).update(planTime=pt, priority=pri)

    except Exception as e:
        return JsonResponse({'code': 1, 'data': str(e)})

    return JsonResponse({'code': 0, 'data': 'succ'})

def kill_clean_task(request, id=0):
    print('<kill_clean_task>')
    if request.method == 'POST':
        data = json.loads(request.POST.get('data'))
        tid = data.get('tid', '')

    try:
        CleanBatchLog.objects.filter(batch_id=id, task_id=tid).update(status='completed',process='killed')
        CleanCron.objects.filter(batch_id=id, task_id=tid).update(emergency=1,completedTime=datetime.datetime(1970, 1, 1, 0, 0))

    except Exception as e:
        return JsonResponse({'code': 1, 'data': str(e)})

    return JsonResponse({'code': 0, 'data': 'succ'})

def save(request):
    if request.method == 'POST':
        # print("eid",request.GET.get('sql'))
        if (request.GET.get('eid')!=None) and (request.GET.get('tb')!=None) and (request.GET.get('sql')!=None):
            eid = request.GET.get('eid')
            tb = request.GET.get('tb')
            sql_list = """{}""".format(request.GET.get('sql'))
            sql = r"""{}""".format(sql_list)
            sql = sql.replace(r'@\n@','\n ')
            print(sql)
            try:
                session = connect_clickhouse.connect(0)
                cursor = session.execute(sql)
                fields = cursor._metadata.keys
                data = pd.DataFrame([dict(zip(fields, item)) for item in cursor.fetchall()])
                data_json = data.to_json(orient='records')
                print(data_json)
                return JsonResponse({'code': 200, "message": "success"})
            except:
                return JsonResponse({'code': 500, "message": "fail"})
        else:
            return JsonResponse({'code': 443,"error":"start_date与end_date不能为空"})
    else:
        return JsonResponse({'code': 403,"error":"当前服务器拒绝GET方式请求，请使用POST方式"})

@csrf_exempt
def search_cleaning_process(request):
    if request.method == 'POST':
        if (request.META.get('HTTP_NAME') == "nint") and (request.META.get('HTTP_PASSWORD') == "cong.xixuan"):
            if (request.POST.get('batch_id')!=None) and (request.POST.get('uuid2')!=None):
                try:
                    import importlib
                    import sys
                    sys.path.append(r'C:\Users\zeng.xiangyan\Desktop\my_sop\my_sop\cleaning\project')
                    batch_id = request.POST.get('batch_id')
                    uuid2 = request.POST.get('uuid2')
                    try:
                        batch = importlib.import_module(f'batch_{batch_id}')
                        result = batch.get_clean_brush_search(uuid2)
                        return JsonResponse({'code': 200, "result": result})
                    except AttributeError:
                        print(f"Module does not have the attribute batch_{batch_id}")

                    # uuid2 = batch174.get_uuid2(start_date,end_date)
                    # uuid2 = [7265910549631285524, 7272587973959240363, 7269618240292168630, 7272928620681602765, 7267767130849207408, 7272587973952759679, 7272304248517184438, 7263285731916779524, 7263285731916781410, 7263639121820285050, 7272587973939311759, 7269251338401511593, 7270725229464701051, 7272213689030278485, 7271105149366931793, 7268138207439924824, 7271105149366911094, 7264719147180120362, 7265534795846479853, 7268833150322955590, 7267353482562588571, 7263639121834163696, 7272304248517028113, 7270725229453136668, 7270362077092685069, 7267767130849028766, 7266971342139576206, 7274442905890867383, 7269618240279223560, 7270725229453156355, 7265910549630971777, 7272587973952757109, 7272304248486631211, 7268462301379917624, 7272304248486956386, 7271105149366568381, 7263639121819979393, 7267767130848740196, 7272587973952753392, 7271105149376409792, 7272304248486661987, 7271105149366566810, 7265097319049534805, 7268138207439569558, 7271105149366543529, 7267353482548442100, 7272304248486944417, 7271105149366579800, 7263639121819934205, 7267767130848690198, 7269618240294681826, 7274797141900404277, 7265534795822135466, 7266229558328425789, 7276349368863535126, 7263285731927962036, 7276349368863539475, 7266971342128856574, 7276349368863539784, 7266971342128876250, 7270725229452718196, 7272304248426527438, 7272928620681316181, 7272213689012845092, 7263285731916602795, 7263990798047168753, 7264719147179791368, 7263990798047165520, 7272304248426565829, 7269991417097665588, 7272304248486521247, 7270362077072613108, 7272304248486605273, 7265534795855139638, 7272304248486506089, 7271105149366531261, 7265534795846374116, 7265097319049531856, 7268833150322559999, 7268462301379908278, 7265097319049531007, 7268833150322554121, 7273669944933404179, 7274030022093120025, 7268833150322554058, 7269618240294680839, 7269251338401186152, 7271105149366531352, 7269618240278793387, 7269251338413738181, 7269251338401688312, 7272304248454142776, 7273331802151606309, 7276349368875821002, 7263639121820422175, 7270362077085163497, 7274030022093438265, 7272304248517641638, 7273669944937776150, 7268462301386074792, 7272304248517632671, 7273669944922951425, 7263639121820424009, 7267767130849207849, 7268138207440091675, 7269618240279372951, 7269251338401685237, 7270362077072968669, 7263285731938906044, 7268462301386074633, 7271105149367070896, 7265534795839580101, 7267795340189102600, 7268138207440091910, 7265534795831244775, 7274030022109639382, 7266592684927991592, 7266229558328947945, 7273032421443814095, 7266592684941739976, 7272928620681295865, 7269251338401168298, 7274442905886284636, 7273669944922473845, 7265534795822116960, 7269618240278789232, 7263990798034580847, 7265534795831000462, 7273331802158237855, 7272213689030069358, 7274030022111308717, 7263990798038652965, 7276349368863507138, 7272928620702204560, 7272304248426506960, 7270725229438834595, 7265910549631168115, 7272304248427014853, 7273669944922611484, 7263285731907118586, 7272304248426986495, 7268833150322764669, 7272304248487605516, 7272304248426982161, 7272304248516505920, 7265534795846370908, 7273331802148278062, 7263285731907027127, 7272587973939176732, 7272928620681303840, 7270725229452943326, 7272304248516480487, 7272213689030188600, 7272213689039978163, 7272304248426981202, 7272304248487601928, 7264355307625073258, 7264719147179979736, 7264355307619159874, 7263285731928026613, 7265534795822178667, 7272213689039978115, 7269618240279246871, 7269991417097769918, 7268462301379958784, 7271420576058582827, 7270725229452938417, 7269251338401276252, 7270725229452822052, 7272304248426803194, 7273331802142372702, 7272304248487335567, 7264355307608394639, 7263285731927993065, 7270362077084945401, 7267353482548503424, 7272213689050781390, 7272587973952918128, 7272213689039926722, 7267353482548494729, 7263285731907078393, 7272928620694587306, 7268833150322662809, 7272304248426799395, 7272213689050781443, 7264355307625064128, 7272304248487266054, 7272213689039930730, 7266592684927568607, 7269991417105904364, 7265910549631036293, 7273331802151485181, 7271420576070764059, 7268462301386061390, 7268833150322653004, 7265910549631053515, 7263285731907081518, 7272304248486461642, 7272304248487071937, 7265910549630991118, 7272213689030125609, 7276349368863556367, 7265910549630999289, 7266592684927519505, 7270725229452781122, 7270725229453353225, 7271105149367098117, 7271420576070758164, 7276349368863556328, 7276349368863557960, 7272587973952767034, 7266971342128902100, 7265910549631433002, 7272304248426782818, 7268833150334354697, 7268833150338740361, 7268833150322625979, 7268462301368849512, 7268138207439594494, 7276349368863545954, 7276349368863547677, 7276349368863548157, 7276349368890813380, 7264355307619116184, 7274797141900408734, 7269251338401242086, 7269251338401233107, 7272304248488367214, 7273669944937773698, 7265910549631382746, 7270362077072957036, 7268138207440061792, 7272587973939422605, 7265910549644177685, 7271420576058835438, 7272304248488378338, 7266971342129323913, 7272304248517548162, 7265534795839552147, 7272304248454085321, 7263285731916846151, 7266592684940199063, 7272304248517537885, 7272304248454081062, 7264355307619249433, 7264719147180262665, 7274797141900630876, 7265097319056935999, 7268462301387666624, 7267353482548886890, 7269251338401654578, 7269618240279343746, 7272304248454081845, 7269251338413848835, 7264355307625098691, 7268138207455002760, 7272304248517714981, 7268462301379956652, 7272304248426970117, 7269618240279012909, 7268462301368886691, 7267767130865974721, 7269251338401355001, 7272304248487383356, 7272304248426972432, 7272304248426974284, 7267767130848797181, 7263639121820097693, 7272928620681460987, 7265910549631102827, 7263639121820101179, 7272304248486499931, 7267353482548511168, 7272304248487365222, 7272691675917900692, 7272304248426814259, 7264355307619131927, 7268462301386061504, 7264355307605555430, 7274797141900627769, 7272304248487567748, 7272304248487567725, 7272304248516440926, 7274797141900627807, 7268833150323135588, 7276349368863571012, 7272304248426971653, 7267353482548620383, 7266592684927667317, 7265910549642748933, 7270725229452957707, 7272304248487730662, 7263990798047227099, 7266229558328606359, 7272213689030214512, 7266592684927665830, 7272304248516618750, 7272304248516580365, 7264719147179992757, 7270725229452957448, 7263285731907119787, 7266229558328599900, 7274797141900553794, 7274797141900555339, 7274797141900555600, 7273669944922676813, 7274797141900556203, 7274797141900555714, 7276349368863627815, 7272213689050783171, 7266971342141322952, 7272304248487654213, 7268462301368991680, 7276349368863631978, 7275256454293654486, 7272587973943422880, 7274797141900555565, 7272304248517498487, 7270725229453263629, 7269251338401627891, 7272304248487796986, 7272304248487780741, 7267353482548644746, 7265910549642750259, 7263285731907149134, 7272304248517460611, 7271105149366794037, 7272304248516678955, 7264719147180036163, 7265534795831113258, 7272304248427083773, 7269618240279081722, 7264719147180262731, 7271105149366792413, 7267767130848941418, 7276349368890818485, 7269618240279088019, 7272304248516650317, 7272691675918461828, 7266229558328650775, 7267353482548648869, 7270725229453287349, 7264719147184216790, 7272304248487784687, 7268462301368992261, 7267767130849168452, 7264355307619172790, 7270362077072669086, 7266592684927553400, 7271773618071210638, 7269618240278903994, 7263639121820020430, 7273669944935489938, 7268462301379929927, 7271773618076901050, 7269251338401255699, 7272304248487244526, 7265097319043521199, 7267353482548483829, 7263990798047189237, 7273331802142369454, 7270725229452809812, 7274797141900617269, 7265097319056836528, 7264355307625063751, 7269251338401260608, 7270725229452811576, 7272304248426791098, 7272304248487239212, 7265534795853528257, 7264355307625063351, 7272213689039925496, 7268833150322646206, 7269991417112478269, 7272213689050781249, 7272213689046497151, 7266592684927551066, 7273669944922726774, 7271420576058688195, 7273331802158329403, 7267353482548681861, 7263285731916732240, 7273331802148286561, 7273331802142379720, 7272304248487835724, 7273669944935513345, 7274797141900576080, 7269618240279171614, 7268138207439821424, 7272587973952836649, 7272304248516776305, 7272587973939254071, 7274797141900617360, 7264355307619147708, 7276349368871499708, 7276349368871499067, 7263990798050273071, 7270362077089911282, 7268138207454946234, 7274797141900534739, 7274797141900613029, 7270362077085058821, 7274797141900614017, 7269618240296258521, 7274797141900615139, 7266971342141327220, 7274797141900613818, 7272304248516301024, 7269251338401311056, 7274797141900484905, 7272304248426863423, 7272587973939118945, 7272213689039942617, 7272304248487385991, 7274797141900426714, 7272213689039945709, 7268462301369118246, 7266971342128985084, 7272304248516334145, 7272304248516298184, 7268833150337022505, 7273669944922614472, 7267353482548538179, 7268833150322702130, 7270362077072694419, 7274797141900522088, 7274797141900526858, 7269618240278978924, 7268462301368920949, 7270725229438870335, 7274797141900516406, 7273331802151498900, 7268833150322717228, 7268833150322701946, 7268462301368914523, 7274030022111312833, 7272304248426907627, 7270725229453349508, 7272304248516440689, 7269251338401355002, 7263285731935982532, 7269991417116357132, 7272304248487067100, 7271105149366601518, 7270725229452783631, 7272304248456713430, 7263285731916864063, 7275199369886570282, 7267353482562631947, 7272928620681747795, 7263285731916861889, 7267353482562630941, 7275199369886572994, 7274797141900624189, 7275199369886575038, 7269251338417289897, 7270362077076316269, 7275199369886571279, 7275199369886571157, 7269991417105913754, 7275199369886570919, 7269251338417290063, 7274797141900626169, 7274797141900635773, 7274797141900639311, 7270725229453355768, 7268462301383380221, 7266592684927444251, 7272304248426381862, 7268462301368718653, 7273669944935488217, 7276349368863480006, 7272304248486232542, 7272304248426413420, 7272304248426361925, 7264719147179709726, 7275256454293395361, 7271773618071172901, 7268138207439490209, 7274030022093057699, 7271105149366435056, 7274030022093059472, 7265910549630855466, 7269991417097631854, 7273669944922419002, 7263285731935812929, 7276349368863322968, 7275299481287560764, 7275299481287563896, 7275299481331526997, 7275299481407256798, 7275299481287567946, 7275299481287566446, 7275299481287550664, 7275299481287551561, 7275299481287566744, 7275299481287558214, 7267484938604118039, 7275299481287565323, 7275299481420676245, 7275299481287563015, 7275299481331539096, 7275299481287556337, 7275299481331528914, 7276112114858087535, 7275299481287566596, 7275299481457854754, 7276112114862147774, 7276112114858087835, 7275299481420673256, 7275299481287559599, 7268438863733505305, 7275299481309722784, 7275299481351407489, 7275299481384841305, 7275299481287562814, 7275299481287556549, 7275299481287536014, 7275299481331513626, 7275299481287538869, 7275299481287531478, 7275299481331518385, 7275299481287536741, 7275299481331516588, 7275299481370352186, 7275299481384837888, 7275299481331517519, 7275299481287525182, 7275299481370352122, 7275299481287525680, 7275299481287529938, 7275299481287542076, 7275299481287528975, 7275299481287539945, 7275299481331511133, 7275299481384839030, 7275299481287528679, 7275299481384838568, 7275299481287539774, 7275299481331513943, 7275299481361573619, 7275299481331514880, 7275299481420670203, 7275299481287531349, 7275299481287529045, 7275299481287537932, 7275299481287528779, 7275299481287518624, 7275299481287513462, 7275299481458900324, 7275299481287517317, 7275299481287519449, 7275299481384832380, 7275299481420663192, 7275299481287523188, 7275299481331502607, 7275299481287510779, 7275299481287509500, 7275299481446917505, 7275299481407254251, 7275299481331505457, 7275299481287517791, 7275299481370347701, 7275299481287519728, 7275299481287521566, 7275299481287521752, 7270713396801572262, 7275299481331508229, 7275299481331508110, 7276112114859947303, 7275299481287512035, 7275299481331502172, 7275299481375125996, 7275299481361570569, 7274359471851529386, 7275299481287516053, 7275299481331506223, 7275299481287502584, 7275299481331490516, 7275299481287502932, 7275299481287502798, 7275299481287506401, 7275299481331487932, 7275299481287504632, 7275299481287505550, 7275299481331492025, 7275299481287506075, 7275299481287508139, 7275299481287501145, 7275299481420660296, 7275299481331487087, 7275299481331492170, 7275299481287507519, 7275299481287502325, 7275299481287508671, 7275299481287503772, 7275299481331492404, 7275299481287502295, 7275299481331488709, 7275299481331488110, 7270713396805479730, 7275299481287505267, 7275299481384829068, 7275299481287508627, 7275299481287507499, 7275299481361568012, 7275299481310898010, 7275299481294822872, 7275299481294821890, 7275299481294828362, 7275299481294818836, 7275299481294816034, 7276112114863916059, 7276112114862415762, 7276112114858447943, 7276112114858447846, 7275299481294829543, 7275299481456824016, 7263216982363099822, 7275299481350511880, 7275299481425667552, 7270328889858059839, 7275299481288820841, 7269282378124554000, 7270713396811480946, 7267338879663724161, 7269944215406185341, 7264694704514064015, 7269944215406185340, 7275299481384847733, 7275299481287586533, 7275299481287585173, 7275299481446920732, 7275299481287590472, 7275299481331555446, 7275299481287660340, 7275299481331553220, 7275299481384849086, 7275299481287588498, 7275299481331548132, 7275299481384848984, 7273736778899522008, 7275299481400766610, 7275299481407257684, 7275299481287584438, 7275299481331547842, 7275299481446920335, 7275299481437216726, 7276112114859950489, 7275299481384847148, 7276112114862148910, 7275299481287590076, 7275299481446920828, 7275299481287491542, 7275299481287590225, 7275299481331554555, 7275299481287580350, 7275299481331550658, 7275299481287584766, 7275299481331542601, 7275299481331545397, 7267716295615668914, 7275299481384845894, 7275299481287576084, 7275299481316016726, 7275299481361578659, 7275299481287577709, 7275299481420678606, 7275299481361578575, 7275299481287576009, 7275299481331547628, 7275299481331543099, 7275299481331546131, 7275299481287577078, 7275299481287572851, 7275299481287576552, 7275299481384846518, 7275299481287573832, 7276112114859949987, 7276112114862148457, 7276112114863637372, 7275299481287576446, 7276112114858088977, 7270085841950340921, 7275299481287572800, 7275299481384846802, 7275299481420678818, 7275299481442872603, 7275299481287576736, 7275299481389352041, 7275299481294824634, 7275299481294827062, 7275299481339078060, 7275299481408469259, 7275299481294826258, 7275299481363083419, 7275299481408468907, 7275299481294825318, 7275299481294826145, 7275299481361563133, 7275299481453392513, 7275299481453392484, 7275299481449245665, 7275299481402110224, 7275299481339078472, 7275299481294827373, 7275299481294826560, 7275299481331462716, 7275299481389352517, 7276112114858084111, 7275299481384819952, 7275299481287460454, 7276112114863227663, 7276112114859943818, 7275299481331464832, 7276112114860349241, 7275299481287459943, 7275299481432158151, 7275299481287464655, 7275299481287453294, 7275299481287470265, 7275299481331452606, 7275299481331452888, 7275299481331459372, 7266211205937098637, 7275299481287467619, 7275299481287467196, 7275299481361565241, 7275299481400753585, 7275299481287450836, 7275299481331472104, 7275299481287459355, 7275299481331469480, 7275299481287455874, 7275299481287458378, 7275299481287460100, 7275299481287463818, 7275299481331462881, 7275299481287466331, 7275299481287457852, 7275299481384820505, 7275299481287459872, 7275299481287468550, 7275299481456312730, 7275299481331453462, 7275299481441604224, 7275299481287466513, 7276112114859943780, 7275299481287473569, 7275299481350272181, 7275299481309608129, 7275299481350138605, 7275299481374841149, 7275299481365247758, 7267716295616367298, 7275299481309721005, 7275299481393826802, 7275299481374840324, 7275299481431403476, 7275299481309722539, 7275299481309721947, 7275299481309608011, 7275299481431403523, 7275299481309720049, 7275299481309607253, 7275299481309608258, 7275299481309607299, 7275299481431331127, 7276112114863775157, 7275299481309607699, 7275299481309608284, 7276112114862788055, 7275299481375124813, 7275299481331478558, 7275299481403576219, 7275299481370356033, 7276112114858932537, 7275299481309610318, 7275299481309720234, 7275299481287493287, 7275299481331478674, 7275299481331482166, 7267484938604118020, 7275299481331481815, 7275299481287498453, 7275299481420656491, 7275299481287487287, 7275299481287495652, 7275299481287494264, 7275299481331485233, 7275299481370342093, 7275299481370342820, 7275299481287496152, 7275299481453393353, 7275299481287493711, 7275299481287496344, 7275299481287493514, 7275299481331479166, 7275299481287491479, 7263685790922706202, 7275299481287491587, 7275299481287496086, 7275299481287492850, 7275299481453393162, 7275299481287494955, 7275299481287492477, 7276112114861463775, 7275299481287495976, 7276112114863228020, 7275299481288817998, 7275299481310791452, 7275299481310783385, 7275299481310786487, 7275299481432157835, 7275299481310793506, 7275299481385787847, 7275299481456305341, 7275299481310785774, 7275299481309969954, 7275299481288817910, 7275299481310795045, 7275299481422032214, 7275299481453805558, 7275299481351317329, 7275299481310794466, 7275299481310790852, 7273056408835196628, 7275299481310789019, 7275299481403759209, 7275299481310788950, 7275299481310789571, 7275299481410036616, 7275299481310785164, 7276112114861666078, 7275299481310792345, 7275299481310789218, 7275299481365452957, 7275299481459586019, 7275299481310782858, 7275299481287484562, 7275299481370340746, 7275299481370339594, 7275299481287486581, 7275299481287484402, 7275299481287480990, 7275299481370340765, 7275299481287484934, 7275299481287479558, 7275299481446916209, 7275299481420655557, 7275299481331476386, 7275299481287484495, 7275299481331475836, 7275299481287480501, 7275299481287494844, 7275299481287491512, 7275299481453392824, 7275299481287491836, 7275299481331478047, 7275299481384827412, 7276112114858084449, 7271014400996161681, 7275299481331474527, 7263955227109605774, 7275299481287482102, 7275299481370342471, 7275299481287485226, 7275299481287486291, 7276112114859945340, 7275299481331569741, 7275299481331584520, 7275299481361584529, 7268911752519181996, 7275299481287605816, 7275299481331584061, 7275299481287612182, 7275299481331574991, 7275299481287606978, 7275299481287606235, 7275299481287607785, 7275299481287600799, 7275299481331558131, 7275299481331581873, 7275299481438387649, 7275299481287602901, 7275299481287601741, 7275299481384851951, 7267111362348716994, 7275299481361583530, 7267111362348724981, 7275299481287610562, 7275299481287610283, 7275299481287599456, 7275299481287644004, 7275299481287603134, 7267111362348717065, 7275299481287602340, 7275299481407262817, 7275299481331568151, 7275299481287624040, 7275299481331604302, 7275299481446923752, 7275299481287639961, 7275299481331613221, 7275299481361587734, 7267111362348717157, 7275299481400765870, 7275299481453400967, 7275299481453400695, 7275299481310698581, 7275299481287638451, 7275299481400766613, 7275299481287639480, 7275299481311079826, 7275299481287643373, 7275299481287629240, 7275299481331607807, 7275299481287636899, 7275299481442875736, 7275299481287650179, 7275299481287656529, 7275299481331601969, 7275299481331609719, 7275299481287646784, 7275299481287649479, 7275299481287636313, 7275299481287657183, 7275299481384860337, 7275299481384863219, 7275299481458390380, 7275299481449665447, 7275299481449665476, 7275299481444379238, 7275299481449665388, 7275299481444379283, 7275299481449665433, 7275299481449246010, 7275299481449665461, 7275299481310900298, 7275299481444379272, 7275299481458390364, 7275299481456824102, 7275299481444379246, 7275299481438996930, 7275299481287656526, 7275299481438387789, 7275299481310899061, 7275299481287602075, 7275299481453401746, 7275299481287653805, 7275299481453401797, 7275299481403769673, 7275299481438996832, 7275299481331453292, 7275299481441188258, 7275299481384863221, 7276112114863882224, 7275299481310904535, 7276112114863638378, 7275299481351924005, 7275299481374758662, 7275299481311361861, 7275299481349940005, 7275299481449050844, 7275299481431219883, 7275299481317599435, 7275299481349957123, 7275299481444205846, 7275299481317598216, 7274359471849263168, 7275299481349958691, 7270328889860468524, 7275299481409827017, 7275299481349945821, 7276112114862776655, 7275299481310840239, 7275299481309448969, 7275299481431221642, 7275299481349944889, 7275299481349957843, 7273240826147382544, 7275299481309449257, 7275299481309449208, 7275299481309449233, 7264528317477291260, 7274032569014449763, 7275299481309445300, 7270095969482501155, 7275299481309438484, 7274848926321087242, 7274848926321087333, 7274848926321087241, 7274848926321087243, 7274848926321087328, 7274848926321087240, 7274848926321087239, 7274848926321087331, 7274848926321087251, 7274848926321087330, 7274848926321087332, 7274848926321087360, 7274848926321087365, 7274848926321087259, 7274848926321087363, 7274848926321087290, 7274848926321087352, 7274848926321087304, 7274848926321087323, 7274848926321087317, 7274848926321087256, 7274848926321087295, 7274848926321087371, 7274848926321087327, 7274848926321087257, 7274848926321087264, 7274848926321087354, 7274848926321087353, 7274848926321087237, 7274848926321087340, 7274848926321087358, 7274848926321087372, 7274848926321087341, 7274848926321087287, 7274848926321087366, 7274848926321087291, 7274848926321087370, 7274848926321087276, 7274848926321087343, 7274848926321087229, 7274848926321087233, 7274848926321087273, 7274848926321087362, 7274848926321087225, 7274848926321087226, 7274848926321087923, 7274848926321087920, 7274848926321087928, 7274848926321087312, 7274848926321087314, 7274848926321087346, 7274848926321087921, 7274848926321087349, 7274848926321087922, 7274848926321087350, 7274848926321087927, 7274848926321087272, 7274848926321087313, 7274848926321087285, 7274848926321087924, 7274848926321086835, 7274848926321086834, 7274848926321087182, 7274848926321087183, 7274848926321087186, 7274848926321087201, 7274848926321087220, 7274848926321087206, 7274848926321087196, 7274848926321087218, 7274848926321087221, 7274848926321087188, 7274848926321087215, 7274848926321087212, 7274848926321087176, 7274848926321087216, 7274848926321087200, 7274848926321087944, 7274848926321087953, 7274848926321087954, 7274848926321087945, 7274848926321087970, 7274848926321087946, 7274848926321087938, 7274848926321087956, 7274848926321087955, 7274848926321087957, 7274848926321087943, 7274848926321087959, 7274848926321087933, 7274848926321087942, 7274848926321087951, 7274848926321087016, 7274848926321086938, 7274848926321086939, 7274848926321086958, 7274848926321086961, 7274848926321087049, 7274848926321086956, 7274848926321087062, 7274848926321086912, 7274848926321086889, 7274848926321086910, 7274848926321087085, 7274848926321086909, 7274848926321086897, 7274848926321087084, 7274848926321086907, 7274848926321087088, 7274848926321086925, 7274848926321086874, 7274848926321086869, 7274848926321087876, 7274848926321086849, 7274848926321086926, 7274848926321086883, 7274848926321086904, 7274848926321087120, 7274848926321087035, 7274848926321087098, 7274848926321086067, 7274848926321086906, 7274848926321087913, 7274848926321087918, 7274848926321087910, 7274848926321087912, 7274848926321087919, 7274848926321087911, 7274848926321087907, 7274848926321087915, 7274848926321087916, 7274848926321087908, 7274848926321087914, 7274848926321087917, 7274848926321087909, 7274848926321087165, 7274848926321087164, 7274848926321087880, 7274848926321087170, 7274848926321087154, 7274848926321087155, 7274848926321087168, 7274848926321088049, 7274848926321087169, 7274848926321087162, 7274848926321087157, 7274848926321087158, 7274848926321088048, 7274848926321087166, 7274848926321087160, 7274848926321083509, 7274848926321083494, 7274848926321083516, 7274848926321083550, 7274848926321083530, 7274848926321083511, 7274848926321083541, 7274848926321087929, 7274848926321083553, 7274848926321083497, 7274848926321083515, 7274848926321085031, 7274848926321083521, 7274848926321083543, 7274848926321086840, 7274848926321087145, 7274848926321087107, 7274848926321087116, 7274848926321087135, 7274848926321087127, 7274848926321087143, 7274848926321087126, 7274848926321087140, 7274848926321087138, 7274848926321087134, 7274848926321087133, 7274848926321087112, 7274848926321087105, 7274848926321087142, 7274848926321087136, 7274848926321082668, 7274848926321082571, 7274848926321082412, 7274848926321082693, 7274848926321082424, 7274848926321086049, 7274848926321082426, 7274848926321085999, 7274848926321082669, 7274848926321082414, 7274848926321082466, 7274848926321086047, 7274848926321082665, 7274848926321082666, 7274848926321082644, 7274848926321087453, 7274848926321087442, 7274848926321087870, 7274848926321087833, 7274848926321087865, 7274848926321087695, 7274848926321087697, 7274848926321087663, 7274848926321087480, 7274848926321087435, 7274848926321087385, 7274848926321087387, 7274848926321087383, 7274848926321087732, 7274848926321087496, 7274848926321087562, 7274848926321087623, 7274848926321087624, 7274848926321087559, 7274848926321087520, 7274848926321087681, 7274848926321087477, 7274848926321087475, 7274848926321087557, 7274848926321087563, 7274848926321087558, 7274848926321087466, 7274848926321087560, 7274848926321087437, 7274848926321087871, 7274848926321087146, 7274848926321087882, 7274848926321087902, 7274848926321087926, 7274848926321083551, 7274848926321087869, 7274848926321087873, 7274848926321087889, 7274848926321087904, 7274848926321087925, 7274848926321087171, 7274848926321087378, 7274848926321086841, 7274848926321087896, 7274848926321087255, 7274848926321087872, 7274832339157501074, 7274832339157510146, 7274832339157500363, 7274832339157496396, 7274832339157501045, 7274832339157522976, 7274832339157522948, 7274832339157516090, 7274832339157495869, 7274832339157496226, 7274832339157498129, 7274832339157500399, 7274832339157521640, 7274832339157494353, 7274832339157495240, 7274832339157495572, 7274832339157510064, 7274832339157498332, 7274832339157504526, 7274832339157507833, 7274832339157500866, 7274832339157507062, 7274832339157509841, 7274832339157514147, 7274832339157509696, 7274832339157498525, 7274832339157496486, 7274832339157496643, 7274832339157497711, 7274832339157515051, 7274832339157521668, 7274832339157499493, 7274832339157498027, 7274832339157511978, 7274832339157494047, 7274832339157523005, 7274832339157520870, 7274832339157500752, 7274832339157504981, 7274832339157524830, 7274832339157524867, 7274832339157496581, 7274832339157524901, 7274832339157521693, 7274832339157504149, 7274832339157508145, 7274832339157501714, 7274832339157510334, 7274832339157513413, 7274832339157502096, 7274832339157515311, 7274832339157503260, 7274832339157501153, 7274832339157515181, 7274832339157522310, 7274832339157510029, 7274832339157510054, 7274832339157497072, 7274832339157497692, 7274832339157510204, 7274832339157468002, 7274832339157495777, 7274832339157495293, 7274832339157515423, 7274832339157505755, 7274832339157498609, 7274832339157505677, 7274832339157505193, 7274832339157515147, 7274832339157501943, 7274832339157515360, 7274832339157501408, 7274832339157519233, 7274832339157519174, 7274832339157503745, 7274832339157526684, 7274832339157504887, 7274832339157504014, 7274832339157504082, 7274832339157504645, 7274832339157513998, 7274832339157498952, 7274832339157503840, 7274832339157499380, 7274832339157513229, 7274832339157505591, 7274832339157496378, 7274832339157513966, 7274832339157520417, 7274832339157505447, 7274832339157520455, 7274832339157504335, 7274832339157499134, 7274832339157509905, 7274832339157509941, 7274832339157480043, 7274832339157495135, 7274832339157500101, 7274832339157504313, 7274832339157494913, 7274832339157497739, 7274832339157505118, 7274832339157499437, 7274832339157502436, 7274832339157499076, 7274832339157522473, 7274832339157494847, 7274832339157494447, 7274832339157494599, 7274832339157516351, 7274832339157522456, 7274832339157494803, 7274832339157516221, 7274832339157494470, 7274832339157525331, 7274832339157516380, 7274832339157516302, 7274832339157494648, 7274832339157525162, 7274832339157524498, 7274832339157514697, 7274832339157512935, 7274832339157495219, 7274832339157495180, 7274832339157517458, 7274832339157524932, 7274832339157507597, 7274832339157513200, 7274832339157527249, 7274832339157505274, 7274832339157522565, 7274832339157515503, 7274832339157495474, 7274832339157503194, 7274832339157509593, 7274832339157523216, 7274832339157507002, 7274832339157495169, 7274832339157496784, 7274832339157508881, 7274832339157498233, 7274832339157509974, 7274832339157508829, 7274832339157511812, 7274832339157523763, 7274832339157496915, 7274832339157510412, 7274832339157524632, 7274832339157498424, 7274832339157518863, 7274832339157519213, 7274832339157500224, 7274832339157495325, 7274832339157465263, 7274832339157465971, 7274832339157465198, 7274832339157465053, 7274832339157446189, 7274832339157495262, 7274832339157497373, 7274832339157472961, 7274832339157481397, 7274832339157507092, 7274832339157460390, 7274832339157488038, 7274832339157461093, 7274832339157503721, 7274832339157499699, 7274832339157513752, 7274832339157519836, 7274832339157519858, 7274832339157522876, 7274832339157519896, 7274832339157522937, 7274832339157512416, 7274832339157522908, 7274832339157521059, 7274832339157498905, 7274832339157525004, 7274832339157519750, 7274832339157524333, 7274832339157431874, 7274832339157443187, 7274832339157434147, 7274832339157434741, 7274832339157439569, 7274832339157433782, 7274832339157435428, 7274832339157442823, 7274832339157437439, 7274832339157433048, 7274832339157442347, 7274832339157439371, 7274832339157432986, 7274832339157440877, 7274832339157441525, 7274832339157493971, 7274832339157521132, 7274832339157500505, 7274832339157497653, 7274832339157499675, 7274832339157515023, 7274832339157521101, 7274832339157520953, 7274832339157494175, 7274832339157505799, 7274832339157494218, 7274832339157521049, 7274832339157523171, 7274832339157516036, 7274832339157519521, 7274832339157508771, 7274832339157493914, 7274832339157510988, 7274832339157510376, 7274832339157508730, 7274832339157512045, 7274832339157514623, 7274832339157511058, 7274832339157524208, 7274832339157496736, 7274832339157522690, 7274832339157516545, 7274832339157494008, 7274832339157521304, 7274832339157497579, 7274832339157515829, 7274832339157441542, 7274832339157510077, 7274832339157515786, 7274832339157436989, 7274832339157523729, 7274832339157510311, 7274832339157514725, 7274832339157443541, 7274832339157518104, 7274832339157523235, 7274832339157510343, 7274832339157482424, 7274832339157482300, 7274832339157442772, 7274848926321082816, 7274848926321085071, 7274848926321085322, 7274848926321082817, 7274848926321082793, 7274848926321085074, 7274848926321082763, 7274848926321082776, 7274848926321082809, 7274848926321082783, 7274848926321082821, 7274848926321085321, 7274848926321082753, 7274848926321082777, 7274848926321082762, 7274848926321083351, 7274848926321085113, 7274848926321083396, 7274848926321083329, 7274848926321083339, 7274848926321083468, 7274848926321085115, 7274848926321083461, 7274848926321085134, 7274848926321083328, 7274848926321083361, 7274848926321083363, 7274848926321083332, 7274848926321083354, 7274848926321083333, 7274848926321085552, 7274848926321085551, 7274848926321085557, 7274848926321085550, 7274848926321085546, 7274848926321085544, 7274848926321085547, 7274848926321085554, 7274848926321085543, 7274848926321083223, 7274848926321083277, 7274848926321083225, 7274848926321085030, 7274848926321083202, 7274848926321083174, 7274848926321083292, 7274848926321083230, 7274848926321083219, 7274848926321083299, 7274848926321083224, 7274848926321083288, 7274848926321083300, 7274848926321083237, 7274848926321083251, 7274848926321085098, 7274848926321085038, 7274848926321083184, 7274848926321085095, 7274848926321082950, 7274848926321083144, 7274848926321082994, 7274848926321085082, 7274848926321083106, 7274848926321085131, 7274848926321083182, 7274848926321083027, 7274848926321083058, 7274848926321083031, 7274848926321082965, 7274848926321083633, 7274848926321083602, 7274848926321085122, 7274848926321083593, 7274848926321085114, 7274848926321083604, 7274848926321083662, 7274848926321083601, 7274848926321083646, 7274848926321085105, 7274848926321083608, 7274848926321083650, 7274848926321083607, 7274848926321083115, 7274848926321085138, 7274848926321082911, 7274848926321082873, 7274848926321082995, 7274848926321082949, 7274848926321082968, 7274848926321085020, 7274848926321082971, 7274848926321082885, 7274848926321082871, 7274848926321082914, 7274848926321085037, 7274848926321082893, 7274848926321082979, 7274848926321082978, 7274848926321085048, 7274848926321083696, 7274848926321083699, 7274848926321085039, 7274848926321083787, 7274848926321083786, 7274848926321083701, 7274848926321083737, 7274848926321085124, 7274848926321083692, 7274848926321083704, 7274848926321083745, 7274848926321085163, 7274848926321083725, 7274848926321083775, 7274848926321083718, 7274848926321085351, 7274848926321084065, 7274848926321084091, 7274848926321084062, 7274848926321084077, 7274848926321084071, 7274848926321084064, 7274848926321084090, 7274848926321084074, 7274848926321084067, 7274848926321084092, 7274848926321086392, 7274848926321085352, 7274848926321086394, 7274848926321085354, 7274848926321085059, 7274848926321085069, 7274848926321085167, 7274848926321085068, 7274848926321085242, 7274848926321085250, 7274848926321085213, 7274848926321085062, 7274848926321085194, 7274848926321085144, 7274848926321083513, 7274848926321085253, 7274848926321085258, 7274848926321085189, 7274848926321085146, 7274848926321083536, 7274848926321083534, 7274848926321085065, 7274848926321083531, 7274848926321083512, 7274848926321083549, 7274848926321083557, 7274848926321083558, 7274848926321083539, 7274848926321083508, 7274848926321083563, 7274848926321083507, 7274848926321083495, 7274848926321083525, 7274848926321083560, 7272304248498539493, 7264719147184710077, 7264719147184678294, 7270362077076316262, 7267795340189598055, 7268833150338920645, 7269997945445171233, 7266971342133365563, 7273331802144892851, 7272304248498541856, 7273331802160132106, 7266229558333376680, 7276349368876627426, 7269251338417553130, 7272304248498350160, 7275199369887669446, 7275199369887677837, 7275199369887678416, 7275199369887670436, 7275199369887668371, 7275199369887678473, 7272304248475743747, 7271420576062183581, 7276349368873051659, 7275199369887669946, 7275199369887672217, 7275199369887668197, 7275199369887679725, 7275199369887678354, 7270362077075952334, 7275199369886597157, 7275199369886596552, 7275199369886597327, 7275199369886597447, 7275199369886596429, 7275199369886597307, 7275199369886590008, 7275199369886589430, 7275199369886596759, 7275199369886599106, 7275199369886589639, 7275256454294378166, 7275199369886597501, 7275199369886597280, 7275199369886596848, 7275199369887673920, 7268462301381113436, 7272304248495462414, 7264719147184217079, 7269991417118432739, 7270362077092710047, 7272304248495464659, 7268833150327731689, 7270362077092710072, 7263990798048540824, 7264719147184227644, 7265534795855336570, 7265534795833426348, 7272587973942784782, 7263990798029993783, 7275199369887712373, 7268462301372817284, 7275199369887699107, 7272304248499463950, 7263285731919043042, 7266229558333317860, 7269991417113902446, 7269251338405635765, 7268462301381278441, 7272213689032562064, 7272213689024426054, 7269251338405431890, 7275199369887695496, 7272304248498239793, 7272304248495702838, 7270725229457631767, 7265534795850534223, 7275199369887661965, 7263639121824071499, 7268462301372197485, 7263285731929550804, 7275199369887664274, 7276349368891504165, 7266971342141612985, 7276349368891504137, 7269618240283110213, 7266229558342535435, 7273331802159954570, 7275199369887661261, 7265534795847717869, 7269991417115891271, 7273331802160126851, 7276349368891574269, 7270725229458208480, 7269618240289657942, 7274797141900642684, 7264719147184660370, 7276349368876593332, 7273669944927269058, 7269251338405424932, 7267767130853237539, 7276349368891571678, 7269991417113844589, 7269251338414948577, 7275256454294430428, 7266229558333621334, 7275199369887721971, 7266971342133588304, 7275199369887722823, 7275199369887722953, 7275199369887722670, 7275199369887723323, 7275199369887723115, 7275256454294432435, 7268833150337302348, 7274797141900650427, 7272304248457413133, 7275199369887722981, 7275199369887723286, 7268462301372918099, 7275199369887746885, 7274797141900651784, 7269618240284089054, 7274797141900652437, 7263990798038847244, 7274797141900652230, 7269618240289664598, 7271420576062767273, 7269991417116734043, 7275199369887747019, 7275199369887747104, 7272304248499376154, 7264355307605933035, 7268833150327970988, 7275199369887746989, 7275199369887766089, 7275199369887769136, 7275199369887776394, 7275199369887704797, 7275199369887772848, 7268833150327750375, 7275199369887762323, 7275199369887663512, 7275199369887771421, 7275199369887715448, 7275199369887753970, 7275199369887756149, 7275199369887772087, 7276349368885144967, 7276349368891571404, 7272304248399014084, 7272304248436876689, 7272304248493399216, 7265534795823340008, 7266229558331434057, 7268138207442421699, 7268138207442430041, 7269618240281822646, 7272304248454322566, 7270725229456095476, 7273331802143873256, 7265534795823340033, 7272304248454322504, 7272304248399014260, 7263639121834350412, 7275299481288684751, 7275299481288853322, 7275299481288661389, 7275299481288849778, 7275299481288851337, 7275299481333112588, 7275299481288663458, 7275299481288661521, 7275299481459218270, 7275299481288849358, 7275299481288861230, 7275299481288850855, 7275299481288672731, 7275299481288849442, 7275299481288660219, 7275299481288783526, 7275299481288783037, 7275299481288779324, 7275299481288800525, 7275299481288796803, 7265258517752184857, 7276112114862189145, 7275299481288771248, 7275299481385777603, 7276112114863645972, 7275299481288796258, 7275299481333059671, 7275299481288808390, 7275299481288783717, 7275299481288800977, 7276112114863848310, 7276112114859623538, 7275299481325362393, 7275299481281830982, 7275299481281831820, 7275299481281831335, 7275299481437156604, 7275299481416981395, 7275299481406400509, 7275299481437562852, 7275299481437156616, 7275299481325363025, 7275299481439566130, 7275299481416982173, 7275299481325363003, 7275299481288783199, 7275299481333056670, 7276112114860011445, 7275299481333059527, 7275299481288796232, 7276112114862190362, 7276112114858144548, 7275299481332762409, 7275299481288785238, 7275299481288775998, 7275299481332762440, 7275299481385778310, 7275299481385777643, 7276112114858144675, 7275299481385813557, 7275299481288896257, 7275299481288734308, 7275299481288736993, 7275299481288747243, 7275299481288731804, 7275299481288742092, 7275299481288730115, 7275299481288736717, 7275299481288744454, 7275299481288736213, 7275299481288736339, 7275299481288729898, 7275299481288737693, 7275299481288733165, 7275299481288749853, 7275299481311958980, 7275299481311959932, 7275299481311959893, 7275299481439039330, 7275299481311959080, 7275299481352512855, 7276112114864173075, 7276112114861012838, 7275299481311959176, 7276112114861665179, 7275299481311962119, 7275299481352515382, 7276112114861678277, 7276112114859045796, 7275299481311960068, 7275299481288705112, 7275299481288715897, 7275299481288703663, 7275299481288706517, 7275299481288703773, 7275299481288716018, 7275299481288702530, 7275299481333137019, 7276112114862186562, 7275299481288721126, 7276112114858142089, 7275299481288707485, 7275299481288707233, 7275299481288702606, 7276112114864104300, 7275299481310456242, 7275299481310452098, 7275299481310456406, 7275299481311699027, 7275299481458000802, 7270367840912354326, 7275299481350911928, 7276112114858976611, 7275299481310456484, 7276112114860936923, 7275299481431809310, 7266327144270169222, 7275299481311709001, 7275299481431806245, 7275299481311703932, 7275299481310180216, 7275299481350684304, 7275299481310977543, 7275299481310179631, 7275299481431699097, 7275299481310983049, 7275299481310177666, 7275299481309833479, 7276112114858959351, 7275299481310982424, 7275299481288896487, 7275299481310179797, 7275299481310177870, 7275299481288749700, 7276112114858959358, 7275299481288921969, 7275299481288729769, 7275299481333138115, 7275299481288924233, 7275299481288922778, 7275299481288915199, 7275299481288919808, 7275299481288736989, 7276112114863646983, 7275299481288925422, 7276112114860019079, 7276112114863647174, 7275299481288925585, 7275299481288922818, 7276112114858151652, 7275299481288828398, 7275299481333087735, 7275299481288824403, 7275299481288825888, 7275299481288823551, 7275299481401085385, 7275299481288828663, 7275299481288824992, 7275299481288825005, 7275299481407571843, 7275299481333092592, 7275299481288828562, 7275299481288828755, 7275299481288825765, 7276112114858146829, 7274832339157468329, 7274832339157463772, 7274832339157447274, 7274832339157469468, 7274832339157483042, 7274832339157486509, 7274832339157461417, 7274832339157459556, 7274832339157445837, 7274832339157468242, 7274832339157462461, 7274832339157467494, 7274832339157462013, 7274832339157449172, 7274832339157445135, 7274832339157471150, 7274832339157471187, 7274832339157482405, 7274832339157446838, 7274832339157446505, 7274832339157468359, 7274832339157489802, 7274832339157482324, 7274832339157445100, 7274832339157451797, 7274832339157451602, 7274832339157453020, 7274832339157456422, 7274832339157480437, 7274832339157458344, 7274832339157532717, 7274832339157532713, 7274832339157528556, 7274832339157446681, 7274832339157487033, 7274832339157475204, 7274832339157476857, 7274832339157486002, 7274832339157464749, 7274832339157453584, 7274832339157459216, 7274832339157447318, 7274832339157458589, 7274832339157448819, 7274832339157460771, 7274832339157447977, 7274832339157444758, 7274832339157464630, 7274832339157491598, 7274832339157445274, 7274832339157446228, 7274832339157467458, 7274832339157480583, 7274832339157471071, 7274832339157470932, 7274832339157446511, 7274832339157471042, 7274832339157451748, 7274832339157463018, 7274832339157460328, 7274832339157463538, 7274832339157482448, 7274832339157462159, 7274832339157491302, 7274832339157474910, 7274832339157445581, 7274832339157446447, 7274832339157457222, 7274832339157467791, 7274832339157446567, 7274832339157471216, 7274832339157446206, 7274832339157474959, 7274832339157464521, 7274832339157485776, 7274832339157467818, 7274832339157465874, 7274832339157459438, 7274832339157468104, 7274832339157475091, 7274832339157491399, 7274832339157445164, 7274832339157475048, 7274832339157469289, 7274832339157446951, 7274832339157490245, 7274832339157460380, 7274832339157460016, 7274832339157479501, 7274832339157446784, 7274832339157445559, 7274832339157456507, 7274832339157469372, 7274832339157451830, 7274832339157454500, 7274832339157448408, 7274832339157453962, 7274832339157477212, 7274832339157446748, 7274832339157481125, 7274832339157466919, 7274832339157461874, 7274832339157478325, 7274832339157466897, 7274832339157490738, 7274832339157458971, 7274832339157466218, 7274832339157469938, 7274832339157471763, 7274832339157465940, 7274832339157459109, 7274832339157451852, 7274832339157476367, 7274832339157461365, 7274832339157471570, 7274832339157450392, 7274832339157464147, 7274832339157454817, 7274832339157471751, 7274832339157447686, 7274832339157471636, 7274832339157452344, 7274832339157454975, 7274832339157474606, 7274832339157474941, 7274832339157471395, 7274832339157477153, 7274832339157486713, 7274832339157448315, 7274832339157467090, 7274832339157479840, 7274832339157476449, 7274832339157468893, 7274832339157477053, 7274832339157474857, 7274832339157486036, 7274832339157466027, 7274832339157486094, 7274832339157488754, 7274832339157471916, 7274832339157466103, 7274832339157484006, 7274832339157451987, 7274832339157472768, 7274832339157488453, 7274832339157483924, 7274832339157456378, 7274832339157473066, 7274832339157464887, 7274832339157477994, 7274832339157473340, 7274832339157472780, 7274832339157483593, 7267353482557329533, 7276349368870097237, 7270725229463587338, 7274030022107068842, 7267353482565210012, 7274797141900657285, 7276349368870095313, 7274797141900657446, 7265534795843082804, 7274797141900657539, 7269618240279395449, 7275199369886578306, 7269991417102524464, 7269991417097937961, 7268138207449114794, 7263990798035034485, 7268833150337524939, 7269251338413858408, 7266971342141931159, 7265534795836193709, 7266592684936834726, 7267353482557334110, 7273669944936448369, 7267353482557334284, 7270725229463591751, 7271105149376532600, 7267353482557334267, 7270725229442437471, 7269251338409756259, 7267353482565211935, 7268462301376548641, 7269618240288521070, 7268462301382559038, 7268462301376548625, 7268833150332597758, 7270362077079977310, 7271105149384935967, 7275256454294485545, 7263285731921245639, 7273032421443794777, 7269251338409756731, 7267795340189557252, 7268462301376541205, 7269251338417784599, 7276349368892453492, 7270362077092389156, 7271105149384933842, 7267767130864768471, 7276349368892455586, 7268833150332586246, 7269251338409754525, 7267767130858016965, 7263990798042441205, 7267767130862678717, 7268462301376540543, 7267353482557329699, 7276349368875837972, 7271420576077648563, 7268833150332584089, 7267353482565209222, 7267767130858016925, 7267353482565210036, 7267353482562633294, 7269991417117036560, 7267767130858016889, 7269618240295297640, 7263285731937960094, 7267353482557329563, 7267353482557329054, 7276349368892454335, 7265910549639905958, 7270362077079982764, 7268462301385013157, 7268833150332596351, 7269618240295298086, 7268833150332596319, 7271773618073794086, 7269251338416107582, 7269991417120217895, 7269618240288519558, 7268138207449103132, 7267767130858011537, 7269991417097936908, 7275256454294635103, 7268833150332583925, 7266971342137929354, 7270362077089204521, 7276349368875837954, 7266971342129380812, 7269618240288506860, 7267353482565209519, 7270362077089205805, 7268833150332474327, 7270725229463583664, 7263285731911112241, 7275256454293734097, 7266971342137824070, 7267353482557230108, 7266971342137933133, 7272213689025953977, 7276112114859312314, 7275299481287710941, 7276112114859958289, 7276112114863230620, 7276112114858095994, 7275299481358110401, 7275299481287711343, 7275299481287711590, 7275299481287711192, 7275299481287710283, 7275299481376803343, 7275299481287672157, 7276112114862153701, 7275299481287716988, 7275299481331638787, 7275299481287699956, 7275299481287701095, 7275299481287716831, 7275299481411022567, 7275299481331638929, 7275299481287671276, 7275299481287671141, 7275299481287671092, 7276112114859957065, 7275299481287714325, 7276112114858095024, 7275299481287681046, 7275299481287714228, 7275299481287681154, 7275299481287684909, 7275299481287691988, 7275299481439833997, 7276112114862153465, 7276112114863230433, 7275299481287692210, 7275299481287691517, 7275299481446924456, 7275299481287691848, 7275299481446924422, 7275299481287692305, 7275299481287697508, 7275299481287697624, 7275299481287698340, 7275299481287697977, 7275299481287697469, 7275299481287698493, 7275299481361591442, 7275299481287698123, 7275299481384870043, 7275299481287697376, 7275299481366842003, 7275299481287715640, 7275299481287715950, 7275299481287674103, 7275299481287699738, 7275299481287716488, 7275299481420702838, 7275299481287670766, 7275299481287716730, 7275299481287716677, 7275299481287703865, 7275299481287704906, 7275299481287706211, 7275299481287705917, 7275299481287705221, 7276112114858095740, 7275299481287706000, 7276112114862153764, 7275299481287703647, 7275299481287705456, 7275299481287708916, 7276112114863856177, 7275299481287709306, 7275299481287699384, 7275299481287709199, 7275299481331624830, 7275299481331636050, 7275299481287684423, 7275299481287672078, 7275299481287679414, 7275299481287693860, 7275299481287693071, 7276112114858095530, 7276112114859957749, 7275299481453402604, 7275299481287693682, 7275299481400768260, 7275299481384869178, 7275299481384869035, 7275299481287692927, 7275299481287714795, 7276112114858096194, 7276112114862153204, 7275299481358110460, 7275299481287701579, 7275299481287714455, 7275299481370371565, 7275299481407267295, 7275299481287714522, 7275299481317329005]
                        return JsonResponse({'code': 200, "result": "uuid查询异常"})
                except:
                    return JsonResponse({'code': 500, "result": "uuid查询异常"})
            else:
                return JsonResponse({'code': 443,"error":"start_date与end_date不能为空"})
        else:
            return JsonResponse({'code': 429, "error": "用户名或密码错误"})
    else:
        return JsonResponse({'code': 403,"error":"当前服务器拒绝GET方式请求，请使用POST方式"})